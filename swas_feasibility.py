#!/usr/bin/env python3
"""
SWAS Thermal Feasibility & Auto-Adjustment Engine
===================================================
Verifies that selected primary and secondary sample coolers are thermally
feasible against available area and utility constraints, and automatically
resolves infeasible cases by adjusting primary outlet temperature or,
as a fallback, sample flowrate.

This module works in concert with swas_heat_duty (duty calculation) and
swas_cooler_screen (candidate screening / zone decomposition).

Key engineering principles
--------------------------
- Steam primaries are area-limited by subcooling.
- The primary should NOT be forced to cool near CW inlet temperature.
- The secondary exists for liquid trim cooling.
- The solver shifts duty between stages rather than failing outright.

Usage
-----
Called from swas_cooler_screen.py after initial screening.  Can also be
invoked standalone for debugging:

    python swas_feasibility.py --config swas_config.yaml
"""

import math
from dataclasses import dataclass, field
from typing import Optional, List

from CoolProp.CoolProp import PropsSI

from swas_heat_duty import (
    calc_point,
    F_to_K, K_to_F,
    psig_to_Pa, W_to_BTU_hr,
    lbm_min_to_kg_s, kg_s_to_lbm_min,
    STEAM_PHASES, SUPERHEATED, SATURATED, SUBCOOLED, SUPERCRITICAL,
    classify,
)
from swas_cooler_screen import (
    decompose_steam_primary_zones,
    lmtd_counterflow,
    ua_from_lmtd,
    area_required,
    select_U_assumed,
    primary_max_duty_BTU_hr,
    ZoneUA,
    CP_WATER_BTU_LBF,
    GAL_PER_MIN_TO_LB_HR,
    LBM_PER_MIN_TO_LBM_HR,
    MIN_Q_BTU_HR,
    MIN_LMTD_F,
)

# ---------------------------------------------------------------------------
# Feasibility result data structures
# ---------------------------------------------------------------------------

@dataclass
class FeasibilityResult:
    """Complete feasibility assessment and auto-adjustment result for one sample point."""
    name: str
    phase: str
    is_steam: bool

    # --- Original conditions ---
    T_in_F: float
    P_psig: float
    T_prim_F_original: float
    T_out_F: float       # final target (77 F)
    flow_lbm_min_original: float

    # --- Primary cooler feasibility ---
    A_primary_available_ft2: float
    A_primary_required_ft2: Optional[float]
    Q_primary_required: float
    Q_primary_utility_max: float
    is_area_ok: bool
    is_utility_ok: bool
    is_primary_feasible: bool

    # --- Adjusted values (after auto-solve) ---
    T_prim_F_adjusted: float
    flow_lbm_min_adjusted: float
    Q_primary_adjusted: float
    Q_secondary_adjusted: float
    A_primary_required_adjusted: Optional[float]

    # --- Secondary cooler feasibility ---
    Q_secondary_required: float
    Q_secondary_utility_max: float
    A_secondary_available_ft2: Optional[float]
    A_secondary_required_ft2: Optional[float]
    is_secondary_feasible: bool
    chw_flow_required_gpm: float

    # --- Utility side temperatures & flows (primary CW) ---
    cw_supply_F: float = 0.0
    cw_return_F: float = 0.0          # actual CW outlet temp for this point
    cw_flow_gpm: float = 0.0          # CW flow to this cooler (GPM)
    cw_dT_design_F: float = 0.0

    # --- Utility side temperatures & flows (secondary CHW) ---
    chw_supply_F: float = 0.0
    chw_return_F: float = 0.0         # actual CHW return temp for this point
    chw_flow_required_gpm_calc: float = 0.0

    # --- Sample temperatures per stage ---
    T_sample_primary_in_F: float = 0.0
    T_sample_primary_out_F: float = 0.0
    T_sample_secondary_in_F: float = 0.0
    T_sample_secondary_out_F: float = 0.0

    # --- T_sat for reference ---
    T_sat_F: Optional[float] = None

    # --- Zone breakdown (after adjustment) ---
    steam_zones_adjusted: Optional[List[ZoneUA]] = None

    # --- Flags ---
    flags: List[str] = field(default_factory=list)

    # --- Totals for system-level rollup ---
    Q_total: float = 0.0
    chiller_tons: float = 0.0


# ---------------------------------------------------------------------------
# Core area calculation (recompute for arbitrary T_primary_out)
# ---------------------------------------------------------------------------

def compute_primary_area_for_T_out(
    T_in_F: float,
    P_psig: float,
    T_primary_out_F: float,
    flow_lbm_min: float,
    phase: str,
    T_sat_F: Optional[float],
    cw_in_F: float,
    cw_max_gpm: float,
    F_correction: float = 1.0,
) -> tuple:
    """
    Compute total required primary cooler area for a given primary outlet
    temperature.  Used by the bisection solver.

    Returns (A_total_ft2, Q_primary_BTU_hr, zones_list, cw_dT_F)
    where A_total_ft2 is None if LMTD is indeterminate.
    """
    P_Pa = psig_to_Pa(P_psig)
    m_dot_kg_s = lbm_min_to_kg_s(flow_lbm_min)
    flow_lbhr = flow_lbm_min * LBM_PER_MIN_TO_LBM_HR

    # Enthalpies
    h_in = _get_enthalpy_inlet(T_in_F, P_psig, phase, T_sat_F)
    h_out = PropsSI("H", "T", F_to_K(T_primary_out_F), "P", P_Pa, "Water")

    Q_prim_W = m_dot_kg_s * (h_in - h_out)
    Q_prim_BTU = W_to_BTU_hr(Q_prim_W)

    if Q_prim_BTU <= MIN_Q_BTU_HR:
        return 0.0, 0.0, [], 0.0

    # CW temperature rise
    cw_m_dot_lbhr = cw_max_gpm * 8.34 * 60.0
    cw_dT = Q_prim_BTU / (cw_m_dot_lbhr * CP_WATER_BTU_LBF) if cw_m_dot_lbhr > 0 else 0.0
    cw_out_F = cw_in_F + cw_dT

    if phase in STEAM_PHASES and T_sat_F is not None:
        # Build a synthetic result dict for decompose_steam_primary_zones
        h_g = PropsSI("H", "P", P_Pa, "Q", 1, "Water")
        h_f = PropsSI("H", "P", P_Pa, "Q", 0, "Water")

        if phase == SUPERHEATED:
            Q_dsh_W = m_dot_kg_s * (h_in - h_g)
            Q_cond_W = m_dot_kg_s * (h_g - h_f)
        elif phase == SATURATED:
            Q_dsh_W = 0.0
            Q_cond_W = m_dot_kg_s * (h_in - h_f)
        else:
            Q_dsh_W = 0.0
            Q_cond_W = 0.0

        Q_dsh_BTU = W_to_BTU_hr(Q_dsh_W)
        Q_cond_BTU = W_to_BTU_hr(Q_cond_W)

        # If T_primary_out is above T_sat, no condensation/subcooling zones
        if T_primary_out_F >= T_sat_F:
            # Only desuperheat zone (partial)
            Q_dsh_BTU = Q_prim_BTU
            Q_cond_BTU = 0.0
            Q_subcool_BTU = 0.0
        else:
            Q_subcool_BTU = Q_prim_BTU - Q_dsh_BTU - Q_cond_BTU

        # Build synthetic result dict
        r_synth = {
            "Q_prim": Q_prim_BTU,
            "Q_dsh": Q_dsh_BTU,
            "Q_cond": Q_cond_BTU,
            "T_in_F": T_in_F,
            "T_sat_F": T_sat_F,
            "T_prim_F": T_primary_out_F,
            "cw_dT": cw_dT,
        }
        zones = decompose_steam_primary_zones(r_synth, cw_in_F, cw_max_gpm, F_correction)
        valid_areas = [z.A_required_ft2 for z in zones if z.A_required_ft2 is not None]
        A_total = sum(valid_areas) if valid_areas else None
        return A_total, Q_prim_BTU, zones, cw_dT

    else:
        # Liquid: single-phase counterflow
        cw_actual_out_F = cw_in_F + cw_dT
        lmtd_prim = lmtd_counterflow(T_in_F, T_primary_out_F, cw_in_F, cw_actual_out_F)
        U_prim = select_U_assumed(is_steam_zone=False, T_sample_in_F=T_in_F)
        ua_prim = ua_from_lmtd(Q_prim_BTU, lmtd_prim, F_correction)
        A_total = area_required(ua_prim, U_prim)
        zones = []
        if lmtd_prim is not None and ua_prim is not None and A_total is not None:
            zones = [ZoneUA(
                zone_name="Liquid",
                Q_BTU_hr=Q_prim_BTU,
                lmtd_F=round(lmtd_prim, 2),
                U_assumed=U_prim,
                UA_required=round(ua_prim, 1),
                A_required_ft2=round(A_total, 3),
            )]
        return A_total, Q_prim_BTU, zones, cw_dT


def _get_enthalpy_inlet(T_in_F, P_psig, phase, T_sat_F):
    """Get inlet enthalpy handling phase correctly."""
    P_Pa = psig_to_Pa(P_psig)
    T_in_K = F_to_K(T_in_F)
    if phase == SATURATED:
        return PropsSI("H", "P", P_Pa, "Q", 1, "Water")
    if phase == SUBCOOLED and T_sat_F is not None and T_in_F >= T_sat_F:
        return PropsSI("H", "P", P_Pa, "Q", 0, "Water")
    return PropsSI("H", "T", T_in_K, "P", P_Pa, "Water")


# ---------------------------------------------------------------------------
# Auto-solve: bisection to find T_primary_out that fits available area
# ---------------------------------------------------------------------------

def solve_primary_outlet_temp(
    T_in_F: float,
    P_psig: float,
    flow_lbm_min: float,
    phase: str,
    T_sat_F: Optional[float],
    A_available_ft2: float,
    cw_in_F: float,
    cw_max_gpm: float,
    cw_dT_design: float,
    F_correction: float = 1.0,
    tol_area_ft2: float = 0.01,
    max_iter: int = 60,
) -> tuple:
    """
    Bisection solver: find T_primary_out such that A_required ≈ A_available.

    Lower bound: cw_in_F + 5 (can't cool below CW approach)
    Upper bound: T_sat_F if steam, else T_in_F - 1

    Returns (T_primary_out_F, A_required, Q_primary_BTU, zones, converged)
    """
    # Bounds
    T_lo = cw_in_F + 5.0
    if phase in STEAM_PHASES and T_sat_F is not None:
        T_hi = T_sat_F - 1.0
    else:
        T_hi = T_in_F - 1.0

    if T_lo >= T_hi:
        # Can't solve — return what we can compute at T_lo
        A, Q, zones, _ = compute_primary_area_for_T_out(
            T_in_F, P_psig, T_lo, flow_lbm_min, phase, T_sat_F,
            cw_in_F, cw_max_gpm, F_correction)
        return T_lo, A, Q, zones, False

    # Check that at T_lo area is achievable (i.e., A at T_lo <= A_available)
    A_lo, Q_lo, z_lo, _ = compute_primary_area_for_T_out(
        T_in_F, P_psig, T_lo, flow_lbm_min, phase, T_sat_F,
        cw_in_F, cw_max_gpm, F_correction)

    if A_lo is not None and A_lo > A_available_ft2:
        # Even at maximum cooling, area insufficient — can't solve by temp alone
        return T_lo, A_lo, Q_lo, z_lo, False

    # At T_hi, area should be small (less cooling needed)
    A_hi, Q_hi, z_hi, _ = compute_primary_area_for_T_out(
        T_in_F, P_psig, T_hi, flow_lbm_min, phase, T_sat_F,
        cw_in_F, cw_max_gpm, F_correction)

    if A_hi is not None and A_hi <= A_available_ft2:
        # Already feasible at T_hi (minimum cooling) — no adjustment needed
        return T_hi, A_hi, Q_hi, z_hi, True

    # Bisection: find T where A_required == A_available
    # NOTE: A_required increases as T_primary_out decreases (more cooling = more area)
    # So we search: lower T_out → more area
    for _ in range(max_iter):
        T_mid = (T_lo + T_hi) / 2.0
        A_mid, Q_mid, z_mid, _ = compute_primary_area_for_T_out(
            T_in_F, P_psig, T_mid, flow_lbm_min, phase, T_sat_F,
            cw_in_F, cw_max_gpm, F_correction)

        if A_mid is None:
            # LMTD indeterminate — move away from problematic region
            T_lo = T_mid
            continue

        if abs(A_mid - A_available_ft2) < tol_area_ft2:
            return T_mid, A_mid, Q_mid, z_mid, True

        if A_mid > A_available_ft2:
            # Too much area required — need higher T_out (less cooling)
            T_lo = T_mid
        else:
            # Area available exceeds required — can cool more
            T_hi = T_mid

    # Return best guess after max iterations
    T_mid = (T_lo + T_hi) / 2.0
    A_mid, Q_mid, z_mid, _ = compute_primary_area_for_T_out(
        T_in_F, P_psig, T_mid, flow_lbm_min, phase, T_sat_F,
        cw_in_F, cw_max_gpm, F_correction)
    return T_mid, A_mid, Q_mid, z_mid, True


# ---------------------------------------------------------------------------
# Auto-solve: bisection to find sample flow that fits secondary area
# ---------------------------------------------------------------------------

def solve_secondary_flow(
    T_in_F: float,
    P_psig: float,
    T_prim_F: float,
    T_out_F: float,
    phase: str,
    T_sat_F: Optional[float],
    flow_max_lbm_min: float,
    A_secondary_available_ft2: float,
    chw_supply_F: float,
    chw_return_max_F: float,
    chw_max_gpm: Optional[float],
    F_correction: float = 1.0,
    tol_area_ft2: float = 0.01,
    max_iter: int = 60,
) -> tuple:
    """
    Bisection solver: find the maximum sample flow (lb/min) such that
    the secondary cooler area requirement <= A_secondary_available.

    Returns (flow_lbm_min, converged)
    """
    P_Pa = psig_to_Pa(P_psig)
    h_prim = PropsSI("H", "T", F_to_K(T_prim_F), "P", P_Pa, "Water")
    h_out = PropsSI("H", "T", F_to_K(T_out_F), "P", P_Pa, "Water")

    def _secondary_area(flow_lbm_min_trial):
        m_dot = lbm_min_to_kg_s(flow_lbm_min_trial)
        Q_sec = W_to_BTU_hr(m_dot * (h_prim - h_out))
        if Q_sec <= MIN_Q_BTU_HR:
            return 0.0, Q_sec
        res = check_secondary_feasibility(
            Q_sec, chw_supply_F, chw_return_max_F, chw_max_gpm,
            A_secondary_available_ft2, T_prim_F, T_out_F,
            flow_lbm_min_trial, P_psig, F_correction)
        return (res["A_required_ft2"] or 0.0), Q_sec

    flow_lo = 0.1   # minimum trial flow (lb/min)
    flow_hi = flow_max_lbm_min

    # Check if already feasible at max flow
    A_hi, _ = _secondary_area(flow_hi)
    if A_hi <= A_secondary_available_ft2:
        return flow_hi, True

    # Check if feasible at minimum flow
    A_lo, _ = _secondary_area(flow_lo)
    if A_lo > A_secondary_available_ft2:
        return flow_lo, False  # can't solve even at minimum

    # Bisection: find max flow where A_req <= A_avail
    for _ in range(max_iter):
        flow_mid = (flow_lo + flow_hi) / 2.0
        A_mid, _ = _secondary_area(flow_mid)

        if A_mid <= A_secondary_available_ft2:
            flow_lo = flow_mid  # confirmed feasible
            if A_secondary_available_ft2 - A_mid < tol_area_ft2:
                return round(flow_lo, 3), True
        else:
            flow_hi = flow_mid

    # Return last confirmed feasible flow
    return round(flow_lo, 3), True


# ---------------------------------------------------------------------------
# Secondary feasibility check
# ---------------------------------------------------------------------------

def check_secondary_feasibility(
    Q_secondary_BTU: float,
    chw_supply_F: float,
    chw_return_max_F: float,
    chw_max_gpm: Optional[float],
    A_secondary_available_ft2: Optional[float],
    T_prim_F: float,
    T_out_F: float,
    flow_lbm_min: float,
    P_psig: float,
    F_correction: float = 1.0,
) -> dict:
    """
    Check if the secondary (chiller) cooler can handle the remaining duty.

    Returns dict with:
        is_feasible, Q_utility_max, chw_flow_required_gpm,
        A_required_ft2, flags
    """
    flags = []

    # CHW utility limit
    if chw_max_gpm is not None:
        chw_m_dot_lbhr = chw_max_gpm * 8.34 * 60.0
        Q_chw_max = chw_m_dot_lbhr * CP_WATER_BTU_LBF * (chw_return_max_F - chw_supply_F)
    else:
        Q_chw_max = float("inf")

    is_utility_ok = Q_secondary_BTU <= Q_chw_max

    # Required CHW flow
    dT_chw = chw_return_max_F - chw_supply_F
    if dT_chw > 0 and Q_secondary_BTU > 0:
        chw_flow_lbhr = Q_secondary_BTU / (CP_WATER_BTU_LBF * dT_chw)
        chw_flow_gpm = chw_flow_lbhr / (8.34 * 60.0)
    else:
        chw_flow_gpm = 0.0

    # Area check (if available area specified)
    A_req = None
    is_area_ok = True
    if A_secondary_available_ft2 is not None and Q_secondary_BTU > MIN_Q_BTU_HR:
        lmtd_sec = lmtd_counterflow(T_prim_F, T_out_F, chw_supply_F, chw_return_max_F)
        U_sec = select_U_assumed(is_steam_zone=False, T_sample_in_F=T_prim_F)
        ua_sec = ua_from_lmtd(Q_secondary_BTU, lmtd_sec, F_correction)
        A_req = area_required(ua_sec, U_sec)
        if A_req is not None:
            is_area_ok = A_req <= A_secondary_available_ft2

    if not is_utility_ok:
        flags.append("Utility Limited (CHW)")
    if not is_area_ok:
        flags.append("Area Limited (secondary)")

    return {
        "is_feasible": is_utility_ok and is_area_ok,
        "Q_utility_max": Q_chw_max,
        "chw_flow_required_gpm": round(chw_flow_gpm, 2),
        "A_required_ft2": round(A_req, 3) if A_req is not None else None,
        "is_utility_ok": is_utility_ok,
        "is_area_ok": is_area_ok,
        "flags": flags,
    }


# ---------------------------------------------------------------------------
# Main feasibility engine — per sample point
# ---------------------------------------------------------------------------

def assess_feasibility(
    r: dict,
    A_primary_available_ft2: float,
    cw_supply_F: float,
    cw_max_gpm: float,
    cw_dT_design: float,
    chw_supply_F: float = 55.0,
    chw_return_max_F: float = 65.0,
    chw_max_gpm: Optional[float] = None,
    A_secondary_available_ft2: Optional[float] = None,
    F_correction: float = 1.0,
    allow_flow_reduction: bool = True,
    max_secondary_flow_lbh: Optional[float] = None,
) -> FeasibilityResult:
    """
    Full feasibility assessment and auto-adjustment for one sample point.

    Steps:
      1. Check if original primary duty fits in available area and utility
      2. If infeasible, solve for adjusted primary outlet temperature
      3. Recalculate duty split
      4. Check secondary feasibility
      5. If still infeasible, reduce sample flowrate (optional fallback)

    Parameters
    ----------
    r : dict
        Result from calc_point() for this sample point.
    A_primary_available_ft2 : float
        Maximum heat-transfer area of the selected primary cooler.
    cw_supply_F, cw_max_gpm, cw_dT_design :
        Cooling water constraints.
    chw_supply_F, chw_return_max_F, chw_max_gpm :
        Chilled water constraints.
    A_secondary_available_ft2 : float or None
        Maximum area of secondary cooler (None to skip area check).
    F_correction : float
        LMTD correction factor.
    allow_flow_reduction : bool
        If True, flowrate reduction is allowed as a last-resort fallback.

    Returns
    -------
    FeasibilityResult with all assessment details and adjusted values.
    """
    is_steam = r["is_steam"]
    phase = r["phase"]
    T_in_F = r["T_in_F"]
    P_psig = r["P_psig"]
    T_prim_F_orig = r["T_prim_F"]
    T_out_F = r["T_out_F"]
    flow_lbm_min = r["flow_slip"]
    T_sat_F = r["T_sat_F"]
    Q_prim_orig = float(r["Q_prim"])
    Q_chill_orig = float(r["Q_chill"])
    Q_total_orig = float(r["Q_total"])

    flags = []

    # --- Utility limit ---
    cw_out_max_F = cw_supply_F + cw_dT_design
    Q_prim_util_max = primary_max_duty_BTU_hr(cw_max_gpm, cw_supply_F, cw_out_max_F)

    # --- Compute original required area ---
    A_orig, Q_orig_check, zones_orig, cw_dT_orig = compute_primary_area_for_T_out(
        T_in_F, P_psig, T_prim_F_orig, flow_lbm_min, phase, T_sat_F,
        cw_supply_F, cw_max_gpm, F_correction)

    is_area_ok = (A_orig is not None and A_orig <= A_primary_available_ft2)
    is_utility_ok = (Q_prim_orig <= Q_prim_util_max)
    is_primary_feasible = is_area_ok and is_utility_ok

    # --- Initialize adjusted values to original ---
    T_prim_adjusted = T_prim_F_orig
    flow_adjusted = flow_lbm_min
    Q_prim_adjusted = Q_prim_orig
    Q_sec_adjusted = Q_chill_orig
    A_adjusted = A_orig
    zones_adjusted = zones_orig

    # =================================================================
    # AUTO-SOLVE STEP 1: Adjust primary outlet temperature
    # =================================================================
    if not is_primary_feasible:
        # Determine which constraint binds
        if not is_area_ok:
            flags.append("Area Limited")
            # Solve for T_primary_out that fits A_available
            T_solved, A_solved, Q_solved, z_solved, converged = solve_primary_outlet_temp(
                T_in_F, P_psig, flow_lbm_min, phase, T_sat_F,
                A_primary_available_ft2, cw_supply_F, cw_max_gpm,
                cw_dT_design, F_correction)

            if converged and A_solved is not None and A_solved <= A_primary_available_ft2 * 1.01:
                T_prim_adjusted = round(T_solved, 1)
                Q_prim_adjusted = round(Q_solved)
                A_adjusted = round(A_solved, 3) if A_solved else A_orig
                zones_adjusted = z_solved
                flags.append(f"Adjusted Primary Outlet Temp: {T_prim_F_orig:.1f} -> {T_prim_adjusted:.1f} F")
            else:
                flags.append("Temp adjustment alone insufficient — flow reduction required")

        if not is_utility_ok:
            flags.append("Utility Limited")
            # Cap primary duty to utility max and back-calculate T_primary_out
            P_Pa = psig_to_Pa(P_psig)
            h_in = _get_enthalpy_inlet(T_in_F, P_psig, phase, T_sat_F)
            m_dot_kg_s = lbm_min_to_kg_s(flow_lbm_min)
            Q_util_W = Q_prim_util_max / 3.412141633  # BTU/hr -> W
            h_prim_util = h_in - Q_util_W / m_dot_kg_s if m_dot_kg_s > 0 else h_in
            T_prim_util_K = PropsSI("T", "H", h_prim_util, "P", P_Pa, "Water")
            T_prim_util_F = K_to_F(T_prim_util_K)

            # Use whichever limit is more restrictive (higher T_out = less primary duty)
            if T_prim_util_F > T_prim_adjusted:
                T_prim_adjusted = round(T_prim_util_F, 1)
                Q_prim_adjusted = round(Q_prim_util_max)
                # Recompute area at this temperature
                A_util, _, z_util, _ = compute_primary_area_for_T_out(
                    T_in_F, P_psig, T_prim_adjusted, flow_lbm_min, phase, T_sat_F,
                    cw_supply_F, cw_max_gpm, F_correction)
                A_adjusted = round(A_util, 3) if A_util is not None else A_adjusted
                zones_adjusted = z_util
                flags.append(f"Adjusted Primary Outlet Temp (utility): "
                             f"{T_prim_F_orig:.1f} -> {T_prim_adjusted:.1f} F")

        # =================================================================
        # AUTO-SOLVE STEP 2: Recalculate duty split
        # =================================================================
        P_Pa = psig_to_Pa(P_psig)
        h_in = _get_enthalpy_inlet(T_in_F, P_psig, phase, T_sat_F)
        h_prim_adj = PropsSI("H", "T", F_to_K(T_prim_adjusted), "P", P_Pa, "Water")
        h_out = PropsSI("H", "T", F_to_K(T_out_F), "P", P_Pa, "Water")
        m_dot_kg_s = lbm_min_to_kg_s(flow_lbm_min)

        Q_prim_adjusted = round(W_to_BTU_hr(m_dot_kg_s * (h_in - h_prim_adj)))
        Q_sec_adjusted = round(W_to_BTU_hr(m_dot_kg_s * (h_prim_adj - h_out)))

    # =================================================================
    # AUTO-SOLVE STEP 3: Secondary cooler check (initial)
    # =================================================================
    sec_result = check_secondary_feasibility(
        Q_sec_adjusted, chw_supply_F, chw_return_max_F, chw_max_gpm,
        A_secondary_available_ft2, T_prim_adjusted, T_out_F,
        flow_adjusted, P_psig, F_correction)

    is_secondary_feasible = sec_result["is_feasible"]

    # =================================================================
    # AUTO-SOLVE STEP 4: Flowrate reduction fallback
    # =================================================================
    if allow_flow_reduction and not (is_primary_feasible or
            (A_adjusted is not None and A_adjusted <= A_primary_available_ft2 * 1.01)):
        if A_adjusted is not None and A_adjusted > 0:
            reduction_factor = A_primary_available_ft2 / A_adjusted
            flow_adjusted = round(flow_lbm_min * reduction_factor, 2)
            flags.append(f"Adjusted Sample Flow: {flow_lbm_min:.2f} -> {flow_adjusted:.2f} lb/min")

            # Recompute at reduced flow
            A_reduced, Q_reduced, z_reduced, _ = compute_primary_area_for_T_out(
                T_in_F, P_psig, T_prim_adjusted, flow_adjusted, phase, T_sat_F,
                cw_supply_F, cw_max_gpm, F_correction)
            A_adjusted = round(A_reduced, 3) if A_reduced is not None else A_adjusted
            zones_adjusted = z_reduced

            # Recalculate duty split at reduced flow
            P_Pa = psig_to_Pa(P_psig)
            h_in = _get_enthalpy_inlet(T_in_F, P_psig, phase, T_sat_F)
            h_prim_adj = PropsSI("H", "T", F_to_K(T_prim_adjusted), "P", P_Pa, "Water")
            h_out = PropsSI("H", "T", F_to_K(T_out_F), "P", P_Pa, "Water")
            m_dot_reduced = lbm_min_to_kg_s(flow_adjusted)
            Q_prim_adjusted = round(W_to_BTU_hr(m_dot_reduced * (h_in - h_prim_adj)))
            Q_sec_adjusted = round(W_to_BTU_hr(m_dot_reduced * (h_prim_adj - h_out)))

            # Re-check secondary at reduced flow
            sec_result = check_secondary_feasibility(
                Q_sec_adjusted, chw_supply_F, chw_return_max_F, chw_max_gpm,
                A_secondary_available_ft2, T_prim_adjusted, T_out_F,
                flow_adjusted, P_psig, F_correction)
            is_secondary_feasible = sec_result["is_feasible"]

    # =================================================================
    # AUTO-SOLVE STEP 5: Secondary flow reduction
    # =================================================================
    if allow_flow_reduction and not is_secondary_feasible:
        flow_before_sec_adj = flow_adjusted

        # 5a — Cap flow to secondary cooler max rated flow
        if max_secondary_flow_lbh is not None:
            max_sec_lbm_min = max_secondary_flow_lbh / LBM_PER_MIN_TO_LBM_HR
            if flow_adjusted > max_sec_lbm_min:
                flow_adjusted = round(max_sec_lbm_min, 3)
                flags.append(f"Secondary flow cap: {flow_before_sec_adj:.2f} -> {flow_adjusted:.2f} lb/min")

        # 5b — Recompute duties at (possibly capped) flow and re-check
        P_Pa = psig_to_Pa(P_psig)
        h_in_5 = _get_enthalpy_inlet(T_in_F, P_psig, phase, T_sat_F)
        h_prim_5 = PropsSI("H", "T", F_to_K(T_prim_adjusted), "P", P_Pa, "Water")
        h_out_5 = PropsSI("H", "T", F_to_K(T_out_F), "P", P_Pa, "Water")
        m_dot_5 = lbm_min_to_kg_s(flow_adjusted)
        Q_prim_adjusted = round(W_to_BTU_hr(m_dot_5 * (h_in_5 - h_prim_5)))
        Q_sec_adjusted = round(W_to_BTU_hr(m_dot_5 * (h_prim_5 - h_out_5)))

        # Recompute primary area at reduced flow
        A_reduced_5, _, z_reduced_5, _ = compute_primary_area_for_T_out(
            T_in_F, P_psig, T_prim_adjusted, flow_adjusted, phase, T_sat_F,
            cw_supply_F, cw_max_gpm, F_correction)
        A_adjusted = round(A_reduced_5, 3) if A_reduced_5 is not None else A_adjusted
        zones_adjusted = z_reduced_5 if z_reduced_5 else zones_adjusted

        sec_result = check_secondary_feasibility(
            Q_sec_adjusted, chw_supply_F, chw_return_max_F, chw_max_gpm,
            A_secondary_available_ft2, T_prim_adjusted, T_out_F,
            flow_adjusted, P_psig, F_correction)
        is_secondary_feasible = sec_result["is_feasible"]

        # 5c — Bisect on flow if still infeasible
        if not is_secondary_feasible and A_secondary_available_ft2 is not None:
            flow_solved, converged = solve_secondary_flow(
                T_in_F, P_psig, T_prim_adjusted, T_out_F, phase, T_sat_F,
                flow_adjusted, A_secondary_available_ft2,
                chw_supply_F, chw_return_max_F, chw_max_gpm, F_correction)
            if converged:
                flow_adjusted = flow_solved
                flags.append(f"Secondary flow solved: {flow_before_sec_adj:.2f} -> {flow_adjusted:.3f} lb/min")

                # Recompute everything at solved flow
                m_dot_solved = lbm_min_to_kg_s(flow_adjusted)
                Q_prim_adjusted = round(W_to_BTU_hr(m_dot_solved * (h_in_5 - h_prim_5)))
                Q_sec_adjusted = round(W_to_BTU_hr(m_dot_solved * (h_prim_5 - h_out_5)))

                A_solved_5, _, z_solved_5, _ = compute_primary_area_for_T_out(
                    T_in_F, P_psig, T_prim_adjusted, flow_adjusted, phase, T_sat_F,
                    cw_supply_F, cw_max_gpm, F_correction)
                A_adjusted = round(A_solved_5, 3) if A_solved_5 is not None else A_adjusted
                zones_adjusted = z_solved_5 if z_solved_5 else zones_adjusted

                sec_result = check_secondary_feasibility(
                    Q_sec_adjusted, chw_supply_F, chw_return_max_F, chw_max_gpm,
                    A_secondary_available_ft2, T_prim_adjusted, T_out_F,
                    flow_adjusted, P_psig, F_correction)
                is_secondary_feasible = sec_result["is_feasible"]
            else:
                flags.append("Secondary flow solver did not converge")

    # Add final secondary flags (after any flow adjustment)
    flags.extend(sec_result["flags"])

    Q_total_adjusted = Q_prim_adjusted + Q_sec_adjusted

    # --- Compute utility-side temperatures and flows ---
    # CW: back-calculate actual CW outlet temp from adjusted primary duty
    cw_m_dot_lbhr = cw_max_gpm * 8.34 * 60.0
    if cw_m_dot_lbhr > 0 and Q_prim_adjusted > 0:
        cw_actual_dT = Q_prim_adjusted / (cw_m_dot_lbhr * CP_WATER_BTU_LBF)
        cw_return_actual = cw_supply_F + cw_actual_dT
        cw_flow_actual_gpm = cw_max_gpm  # full CW flow assumed
    else:
        cw_return_actual = cw_supply_F
        cw_flow_actual_gpm = 0.0

    # CHW: back-calculate actual CHW return temp
    chw_dT_range = chw_return_max_F - chw_supply_F
    if chw_dT_range > 0 and Q_sec_adjusted > 0:
        chw_flow_lbhr_needed = Q_sec_adjusted / (CP_WATER_BTU_LBF * chw_dT_range)
        chw_flow_gpm_actual = chw_flow_lbhr_needed / (8.34 * 60.0)
        chw_return_actual = chw_return_max_F  # design return at full dT
    else:
        chw_flow_gpm_actual = 0.0
        chw_return_actual = chw_supply_F

    return FeasibilityResult(
        name=r["name"],
        phase=phase,
        is_steam=is_steam,
        T_in_F=T_in_F,
        P_psig=P_psig,
        T_prim_F_original=T_prim_F_orig,
        T_out_F=T_out_F,
        flow_lbm_min_original=flow_lbm_min,
        A_primary_available_ft2=A_primary_available_ft2,
        A_primary_required_ft2=A_orig,
        Q_primary_required=Q_prim_orig,
        Q_primary_utility_max=Q_prim_util_max,
        is_area_ok=is_area_ok,
        is_utility_ok=is_utility_ok,
        is_primary_feasible=is_primary_feasible,
        T_prim_F_adjusted=T_prim_adjusted,
        flow_lbm_min_adjusted=flow_adjusted,
        Q_primary_adjusted=Q_prim_adjusted,
        Q_secondary_adjusted=Q_sec_adjusted,
        A_primary_required_adjusted=A_adjusted,
        Q_secondary_required=Q_sec_adjusted,
        Q_secondary_utility_max=sec_result["Q_utility_max"],
        A_secondary_available_ft2=A_secondary_available_ft2,
        A_secondary_required_ft2=sec_result["A_required_ft2"],
        is_secondary_feasible=is_secondary_feasible,
        chw_flow_required_gpm=sec_result["chw_flow_required_gpm"],
        # Utility-side temperatures & flows
        cw_supply_F=cw_supply_F,
        cw_return_F=round(cw_return_actual, 1),
        cw_flow_gpm=round(cw_flow_actual_gpm, 2),
        cw_dT_design_F=cw_dT_design,
        chw_supply_F=chw_supply_F,
        chw_return_F=round(chw_return_actual, 1),
        chw_flow_required_gpm_calc=round(chw_flow_gpm_actual, 2),
        # Sample side per stage
        T_sample_primary_in_F=T_in_F,
        T_sample_primary_out_F=T_prim_adjusted,
        T_sample_secondary_in_F=T_prim_adjusted,
        T_sample_secondary_out_F=T_out_F,
        T_sat_F=T_sat_F,
        steam_zones_adjusted=zones_adjusted if zones_adjusted else None,
        flags=flags,
        Q_total=Q_total_adjusted,
        chiller_tons=Q_sec_adjusted / 12000.0 if Q_sec_adjusted > 0 else 0.0,
    )


# ---------------------------------------------------------------------------
# Batch feasibility assessment
# ---------------------------------------------------------------------------

def assess_all_points(
    results: list,
    candidates: list,
    scr_cfg: dict,
    cw_supply_F: float,
    cw_max_gpm: float,
    cw_dT_design: float,
) -> list:
    """
    Run feasibility assessment on all sample points.

    For each point, the best-fit (largest area) passing primary candidate is
    used as A_primary_available.  If no candidates passed screening, the
    largest available candidate for that service type is used.
    """
    chw_supply_F = float(scr_cfg.get("chiller_supply_F", 55.0))
    chw_return_F = float(scr_cfg.get("chiller_return_F", 65.0))
    chw_max_gpm = scr_cfg.get("chw_max_flow_gpm", None)
    if chw_max_gpm is not None:
        chw_max_gpm = float(chw_max_gpm)
    F_correction = float(scr_cfg.get("lmtd_correction_F", 1.0))
    allow_flow_reduction = bool(scr_cfg.get("allow_flow_reduction", True))

    feasibility_results = []

    for r in results:
        is_steam = r["is_steam"]

        # Find best available primary area from candidates
        A_primary = _best_available_area(candidates, is_steam, r["P_psig"], r["T_in_F"])

        # Find best available secondary area and max flow
        A_secondary, max_sec_flow_lbh = _best_available_cooler_info(
            candidates, False, r["P_psig"], r["T_prim_F"])

        fr = assess_feasibility(
            r,
            A_primary_available_ft2=A_primary,
            cw_supply_F=cw_supply_F,
            cw_max_gpm=cw_max_gpm,
            cw_dT_design=cw_dT_design,
            chw_supply_F=chw_supply_F,
            chw_return_max_F=chw_return_F,
            chw_max_gpm=chw_max_gpm,
            A_secondary_available_ft2=A_secondary,
            F_correction=F_correction,
            allow_flow_reduction=allow_flow_reduction,
            max_secondary_flow_lbh=max_sec_flow_lbh,
        )
        feasibility_results.append(fr)

    return feasibility_results


def _best_available_area(candidates, is_steam_service, P_psig, T_F):
    """Return the largest available area from compatible candidates."""
    best = 0.0
    for c in candidates:
        # Service type compatibility
        if is_steam_service and c.service_type == "liquid":
            continue
        if not is_steam_service and c.service_type == "steam":
            continue
        # Pressure check
        if c.max_pressure_psig is not None and P_psig > c.max_pressure_psig:
            continue
        # Temperature check
        if c.max_temp_F is not None and T_F > c.max_temp_F:
            continue
        if c.heat_transfer_area_ft2 is not None and c.heat_transfer_area_ft2 > best:
            best = c.heat_transfer_area_ft2
    return best if best > 0 else 2.5  # fallback default


def _best_available_cooler_info(candidates, is_steam_service, P_psig, T_F):
    """Return (area_ft2, max_sample_flow_lbh) for the best-area compatible candidate."""
    best_area = 0.0
    best_max_flow = None
    for c in candidates:
        if is_steam_service and c.service_type == "liquid":
            continue
        if not is_steam_service and c.service_type == "steam":
            continue
        if c.max_pressure_psig is not None and P_psig > c.max_pressure_psig:
            continue
        if c.max_temp_F is not None and T_F > c.max_temp_F:
            continue
        if c.heat_transfer_area_ft2 is not None and c.heat_transfer_area_ft2 > best_area:
            best_area = c.heat_transfer_area_ft2
            best_max_flow = c.max_sample_flow_lbh
    if best_area <= 0:
        return 2.5, None
    return best_area, best_max_flow


# ---------------------------------------------------------------------------
# Feasibility report output
# ---------------------------------------------------------------------------

def print_feasibility_report(feas_results: list) -> None:
    """Print a structured feasibility report to stdout."""
    W = 120

    total_Q_primary = sum(fr.Q_primary_adjusted for fr in feas_results)
    total_Q_secondary = sum(fr.Q_secondary_adjusted for fr in feas_results)
    total_Q = sum(fr.Q_total for fr in feas_results)
    total_tons = sum(fr.chiller_tons for fr in feas_results)
    total_cw = total_Q_primary
    infeasible_count = sum(1 for fr in feas_results if not fr.is_primary_feasible)
    adjusted_count = sum(1 for fr in feas_results
                         if fr.T_prim_F_adjusted != fr.T_prim_F_original)
    flow_adjusted_count = sum(1 for fr in feas_results
                              if fr.flow_lbm_min_adjusted != fr.flow_lbm_min_original)

    print()
    print("=" * W)
    print("  SWAS THERMAL FEASIBILITY & AUTO-ADJUSTMENT REPORT")
    print("=" * W)
    print(f"  Sample points             : {len(feas_results):>6d}")
    print(f"  Originally infeasible     : {infeasible_count:>6d}")
    print(f"  Primary T_out adjusted    : {adjusted_count:>6d}")
    print(f"  Sample flow adjusted      : {flow_adjusted_count:>6d}")
    print(f"  Total CW duty             : {total_cw:>12,.0f} Btu/hr")
    print(f"  Total CHW duty            : {total_Q_secondary:>12,.0f} Btu/hr")
    print(f"  Total chiller load        : {total_tons:>12.2f} tons")
    print(f"  Total combined duty       : {total_Q:>12,.0f} Btu/hr")
    print("=" * W)

    for fr in feas_results:
        print()
        print("-" * W)
        print(f"  SAMPLE POINT : {fr.name}   [{fr.phase}]")
        print(f"  Conditions   : T_in={fr.T_in_F:.0f} F   P={fr.P_psig:.0f} psig")
        print()

        # Primary feasibility
        prim_status = "PASS" if fr.is_primary_feasible else "FAIL -> AUTO-ADJUSTED"
        print(f"  PRIMARY COOLER FEASIBILITY: {prim_status}")
        print(f"    Area    : required = {_fmt_f(fr.A_primary_required_ft2, 3)} ft²"
              f"   available = {fr.A_primary_available_ft2:.3f} ft²"
              f"   {'OK' if fr.is_area_ok else 'FAIL'}")
        print(f"    Utility : Q_req = {fr.Q_primary_required:,.0f} Btu/hr"
              f"   Q_max_CW = {fr.Q_primary_utility_max:,.0f} Btu/hr"
              f"   {'OK' if fr.is_utility_ok else 'FAIL'}")

        if fr.T_prim_F_adjusted != fr.T_prim_F_original:
            print(f"    ADJUSTED: T_primary_out = {fr.T_prim_F_original:.1f} -> "
                  f"{fr.T_prim_F_adjusted:.1f} F")
            print(f"    ADJUSTED: A_required = {_fmt_f(fr.A_primary_required_adjusted, 3)} ft²")

        if fr.flow_lbm_min_adjusted != fr.flow_lbm_min_original:
            print(f"    ADJUSTED: flow = {fr.flow_lbm_min_original:.2f} -> "
                  f"{fr.flow_lbm_min_adjusted:.2f} lb/min")

        # Duty split
        print()
        print(f"    Final duty split:")
        print(f"      Primary (CW)   : {fr.Q_primary_adjusted:>12,.0f} Btu/hr")
        print(f"      Secondary (CHW): {fr.Q_secondary_adjusted:>12,.0f} Btu/hr  "
              f"({fr.chiller_tons:.2f} tons)")
        print(f"      Total          : {fr.Q_total:>12,.0f} Btu/hr")

        # Steam zone breakdown after adjustment
        if fr.steam_zones_adjusted:
            print()
            print(f"    Steam zone breakdown (adjusted):")
            hdr = f"      {'Zone':<14}  {'Q (Btu/hr)':>12}  {'LMTD (F)':>9}  "
            hdr += f"{'U':>6}  {'A_req (ft²)':>11}"
            print(hdr)
            print("      " + "-" * (len(hdr) - 6))
            for z in fr.steam_zones_adjusted:
                row = (f"      {z.zone_name:<14}  {_fmt_c(z.Q_BTU_hr):>12}  "
                       f"{_fmt_f(z.lmtd_F, 1):>9}  "
                       f"{z.U_assumed:>6.0f}  {_fmt_f(z.A_required_ft2, 3):>11}")
                print(row)

        # Secondary feasibility
        print()
        sec_status = "PASS" if fr.is_secondary_feasible else "FAIL"
        print(f"  SECONDARY COOLER FEASIBILITY: {sec_status}")
        print(f"    Q_required     = {fr.Q_secondary_required:>12,.0f} Btu/hr")
        if fr.Q_secondary_utility_max != float("inf"):
            print(f"    Q_utility_max  = {fr.Q_secondary_utility_max:>12,.0f} Btu/hr")
        print(f"    CHW flow req'd = {fr.chw_flow_required_gpm:.2f} GPM")
        if fr.A_secondary_required_ft2 is not None:
            print(f"    A_required     = {fr.A_secondary_required_ft2:.3f} ft²")
        if fr.A_secondary_available_ft2 is not None:
            print(f"    A_available    = {fr.A_secondary_available_ft2:.3f} ft²")

        # Flags
        if fr.flags:
            print()
            print(f"  FLAGS:")
            for f in fr.flags:
                print(f"    - {f}")

    # System totals
    print()
    print("=" * W)
    print(f"  SYSTEM TOTALS (after auto-adjustment)")
    print(f"  Total CW duty              : {total_cw:>12,.0f} Btu/hr")
    print(f"  Total CHW duty             : {total_Q_secondary:>12,.0f} Btu/hr")
    print(f"  Total chiller load         : {total_tons:>12.2f} tons")
    print(f"  Total combined duty        : {total_Q:>12,.0f} Btu/hr")
    print("=" * W)
    print()


def _fmt_f(val, decimals=1) -> str:
    return f"{val:.{decimals}f}" if val is not None else "N/A"


def _fmt_c(val, decimals=0) -> str:
    return f"{val:,.{decimals}f}" if val is not None else "N/A"


# ---------------------------------------------------------------------------
# CSV export for feasibility results
# ---------------------------------------------------------------------------

def export_feasibility_csv(feas_results: list, output_path: str) -> None:
    """Export feasibility results to CSV."""
    import csv
    from pathlib import Path

    fields = [
        "Sample Point", "Phase", "Is Steam",
        "T_in (F)", "P (psig)",
        "T_prim Original (F)", "T_prim Adjusted (F)",
        "Flow Original (lb/min)", "Flow Adjusted (lb/min)",
        "A_primary Required (ft2)", "A_primary Available (ft2)",
        "A_primary Adjusted (ft2)",
        "Q_primary Required (Btu/hr)", "Q_primary Utility Max (Btu/hr)",
        "Q_primary Adjusted (Btu/hr)", "Q_secondary Adjusted (Btu/hr)",
        "Q_total (Btu/hr)", "Chiller Tons",
        "Area OK", "Utility OK", "Primary Feasible",
        "Secondary Feasible", "CHW Flow Req (GPM)",
        "Flags",
    ]

    with open(output_path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields)
        w.writeheader()
        for fr in feas_results:
            w.writerow({
                "Sample Point": fr.name,
                "Phase": fr.phase,
                "Is Steam": str(fr.is_steam),
                "T_in (F)": fr.T_in_F,
                "P (psig)": fr.P_psig,
                "T_prim Original (F)": fr.T_prim_F_original,
                "T_prim Adjusted (F)": fr.T_prim_F_adjusted,
                "Flow Original (lb/min)": fr.flow_lbm_min_original,
                "Flow Adjusted (lb/min)": fr.flow_lbm_min_adjusted,
                "A_primary Required (ft2)": _fmt_f(fr.A_primary_required_ft2, 3),
                "A_primary Available (ft2)": f"{fr.A_primary_available_ft2:.3f}",
                "A_primary Adjusted (ft2)": _fmt_f(fr.A_primary_required_adjusted, 3),
                "Q_primary Required (Btu/hr)": round(fr.Q_primary_required),
                "Q_primary Utility Max (Btu/hr)": round(fr.Q_primary_utility_max),
                "Q_primary Adjusted (Btu/hr)": round(fr.Q_primary_adjusted),
                "Q_secondary Adjusted (Btu/hr)": round(fr.Q_secondary_adjusted),
                "Q_total (Btu/hr)": round(fr.Q_total),
                "Chiller Tons": round(fr.chiller_tons, 2),
                "Area OK": str(fr.is_area_ok),
                "Utility OK": str(fr.is_utility_ok),
                "Primary Feasible": str(fr.is_primary_feasible),
                "Secondary Feasible": str(fr.is_secondary_feasible),
                "CHW Flow Req (GPM)": fr.chw_flow_required_gpm,
                "Flags": " | ".join(fr.flags),
            })

        # Totals row
        w.writerow({
            "Sample Point": "TOTAL",
            "Q_primary Adjusted (Btu/hr)": sum(fr.Q_primary_adjusted for fr in feas_results),
            "Q_secondary Adjusted (Btu/hr)": sum(fr.Q_secondary_adjusted for fr in feas_results),
            "Q_total (Btu/hr)": sum(fr.Q_total for fr in feas_results),
            "Chiller Tons": round(sum(fr.chiller_tons for fr in feas_results), 2),
        })

    print(f"  Feasibility results exported to: {Path(output_path).resolve()}")
    print()


# ---------------------------------------------------------------------------
# Excel report export (openpyxl)
# ---------------------------------------------------------------------------

def export_feasibility_excel(feas_results: list, output_path: str,
                             cfg: dict = None) -> None:
    """
    Export a multi-sheet, formatted Excel workbook with:
      Sheet 1 - Summary:  system-level totals and design basis
      Sheet 2 - Point Detail:  one row per sample point, all thermal data
      Sheet 3 - Steam Zones:  zone-by-zone breakdown for steam points
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    from pathlib import Path

    wb = Workbook()

    # ---- Styles ----
    hdr_font   = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    hdr_fill   = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    sub_font   = Font(name="Calibri", bold=True, size=11)
    sub_fill   = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    pass_fill  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill  = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill  = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    tot_font   = Font(name="Calibri", bold=True, size=11)
    tot_fill   = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"))
    center     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    right_al   = Alignment(horizontal="right", vertical="center")
    left_al    = Alignment(horizontal="left", vertical="center", wrap_text=True)

    num_0  = "#,##0"
    num_1  = "#,##0.0"
    num_2  = "#,##0.00"
    num_3  = "#,##0.000"

    def _style_header(ws, row, max_col):
        for col in range(1, max_col + 1):
            c = ws.cell(row=row, column=col)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = center
            c.border = thin_border

    def _style_row(ws, row, max_col, fill=None):
        for col in range(1, max_col + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            if fill:
                c.fill = fill

    def _auto_width(ws, max_col, min_width=10, max_width=30):
        for col in range(1, max_col + 1):
            letter = get_column_letter(col)
            best = min_width
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
                for cell in row:
                    if cell.value is not None:
                        txt_len = len(str(cell.value))
                        if txt_len > best:
                            best = txt_len
            ws.column_dimensions[letter].width = min(best + 3, max_width)

    # ==================================================================
    # SHEET 1 — SUMMARY
    # ==================================================================
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.sheet_properties.tabColor = "2F5496"

    total_Q_primary   = sum(fr.Q_primary_adjusted for fr in feas_results)
    total_Q_secondary = sum(fr.Q_secondary_adjusted for fr in feas_results)
    total_Q           = sum(fr.Q_total for fr in feas_results)
    total_tons        = sum(fr.chiller_tons for fr in feas_results)
    infeasible_count  = sum(1 for fr in feas_results if not fr.is_primary_feasible)
    t_adj_count       = sum(1 for fr in feas_results
                            if fr.T_prim_F_adjusted != fr.T_prim_F_original)
    f_adj_count       = sum(1 for fr in feas_results
                            if fr.flow_lbm_min_adjusted != fr.flow_lbm_min_original)

    # Title
    ws1.merge_cells("A1:D1")
    c = ws1["A1"]
    c.value = "SWAS Thermal Feasibility & Auto-Adjustment Report"
    c.font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    # Design basis
    row = 3
    basis = [
        ("Design Basis", ""),
    ]
    if cfg:
        scr = cfg.get("screening", {})
        basis += [
            ("CW Supply Temp", f"{cfg.get('cw_supply_temp_F', '?')} F"),
            ("CW Max Flow (per cooler)", f"{cfg.get('cw_max_flow_gpm', '?')} GPM"),
            ("CW Design dT", f"{cfg.get('cw_design_delta_T_F', '?')} F"),
            ("CHW Supply Temp", f"{scr.get('chiller_supply_F', '?')} F"),
            ("CHW Return Temp", f"{scr.get('chiller_return_F', '?')} F"),
            ("Target Sample Outlet", f"{cfg.get('target_temp_F', '?')} F"),
            ("LMTD Correction (F)", f"{scr.get('lmtd_correction_F', 1.0)}"),
        ]
    basis += [
        ("", ""),
        ("System Results", ""),
        ("Total Sample Points", len(feas_results)),
        ("Originally Infeasible", infeasible_count),
        ("Primary T_out Adjusted", t_adj_count),
        ("Sample Flow Adjusted", f_adj_count),
        ("Total CW Duty (Btu/hr)", f"{total_Q_primary:,.0f}"),
        ("Total CHW Duty (Btu/hr)", f"{total_Q_secondary:,.0f}"),
        ("Total Chiller Load (tons)", f"{total_tons:.2f}"),
        ("Total Combined Duty (Btu/hr)", f"{total_Q:,.0f}"),
    ]
    for label, val in basis:
        ws1.cell(row=row, column=1, value=label).font = sub_font if val == "" else Font(name="Calibri")
        ws1.cell(row=row, column=2, value=val)
        if val == "":
            ws1.cell(row=row, column=1).fill = sub_fill
            ws1.cell(row=row, column=2).fill = sub_fill
        row += 1

    ws1.column_dimensions["A"].width = 30
    ws1.column_dimensions["B"].width = 22

    # ==================================================================
    # SHEET 2 — POINT DETAIL
    # ==================================================================
    ws2 = wb.create_sheet("Point Detail")
    ws2.sheet_properties.tabColor = "548235"

    headers = [
        # --- Identity ---
        "Sample Point",
        "Phase",
        "T_sat (F)",
        "P (psig)",
        # --- Sample side: Primary Stage ---
        "Sample T_in\nPrimary (F)",
        "Sample T_out\nPrimary Orig (F)",
        "Sample T_out\nPrimary Adj (F)",
        # --- Sample side: Secondary Stage ---
        "Sample T_in\nSecondary (F)",
        "Sample T_out\nFinal (F)",
        # --- Sample flow ---
        "Flow Original\n(lb/min)",
        "Flow Adjusted\n(lb/min)",
        "Flow Original\n(lb/hr)",
        "Flow Adjusted\n(lb/hr)",
        # --- CW Utility ---
        "CW Supply (F)",
        "CW Return (F)",
        "CW Flow (GPM)",
        "CW dT Design (F)",
        # --- CHW Utility ---
        "CHW Supply (F)",
        "CHW Return (F)",
        "CHW Flow Req (GPM)",
        # --- Duties ---
        "Q_primary Req\n(Btu/hr)",
        "Q_primary Adj\n(Btu/hr)",
        "Q_primary Util Max\n(Btu/hr)",
        "Q_secondary Adj\n(Btu/hr)",
        "Q_total\n(Btu/hr)",
        "Chiller Tons",
        # --- Area ---
        "A_prim Required\n(ft2)",
        "A_prim Available\n(ft2)",
        "A_prim Adjusted\n(ft2)",
        "A_sec Required\n(ft2)",
        "A_sec Available\n(ft2)",
        # --- Checks ---
        "Area OK",
        "Utility OK",
        "Primary\nFeasible",
        "Secondary\nFeasible",
        # --- Flags ---
        "Flags / Notes",
    ]

    ncols = len(headers)
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, ncols)
    ws2.row_dimensions[1].height = 40

    for i, fr in enumerate(feas_results, 2):
        row_data = [
            fr.name,
            fr.phase,
            fr.T_sat_F,
            fr.P_psig,
            # Sample temps
            fr.T_sample_primary_in_F,
            fr.T_prim_F_original,
            fr.T_prim_F_adjusted,
            fr.T_sample_secondary_in_F,
            fr.T_sample_secondary_out_F,
            # Flow
            fr.flow_lbm_min_original,
            fr.flow_lbm_min_adjusted,
            round(fr.flow_lbm_min_original * 60, 1),
            round(fr.flow_lbm_min_adjusted * 60, 1),
            # CW utility
            fr.cw_supply_F,
            fr.cw_return_F,
            fr.cw_flow_gpm,
            fr.cw_dT_design_F,
            # CHW utility
            fr.chw_supply_F,
            fr.chw_return_F,
            fr.chw_flow_required_gpm_calc,
            # Duties
            fr.Q_primary_required,
            fr.Q_primary_adjusted,
            fr.Q_primary_utility_max,
            fr.Q_secondary_adjusted,
            fr.Q_total,
            fr.chiller_tons,
            # Area
            fr.A_primary_required_ft2,
            fr.A_primary_available_ft2,
            fr.A_primary_required_adjusted,
            fr.A_secondary_required_ft2,
            fr.A_secondary_available_ft2,
            # Checks
            "PASS" if fr.is_area_ok else "FAIL",
            "PASS" if fr.is_utility_ok else "FAIL",
            "PASS" if fr.is_primary_feasible else "FAIL",
            "PASS" if fr.is_secondary_feasible else "FAIL",
            # Flags
            " | ".join(fr.flags) if fr.flags else "",
        ]
        for col, val in enumerate(row_data, 1):
            c = ws2.cell(row=i, column=col, value=val)
            c.border = thin_border

        # Conditional formatting for PASS/FAIL cells
        for col_idx in range(ncols - 4, ncols):  # last 5 columns (4 check cols + flags)
            c = ws2.cell(row=i, column=col_idx + 1)
            if c.value == "PASS":
                c.fill = pass_fill
                c.alignment = center
            elif c.value == "FAIL":
                c.fill = fail_fill
                c.alignment = center

        # Highlight adjusted flow rows
        if fr.flow_lbm_min_adjusted != fr.flow_lbm_min_original:
            ws2.cell(row=i, column=11).fill = warn_fill  # Flow Adjusted col
            ws2.cell(row=i, column=13).fill = warn_fill  # Flow Adjusted lb/hr
        if fr.T_prim_F_adjusted != fr.T_prim_F_original:
            ws2.cell(row=i, column=7).fill = warn_fill   # T_out Adjusted col

    # Number formats for numeric columns
    fmt_map = {
        # col_index (1-based): format
        3: num_1, 4: num_0,                        # T_sat, P
        5: num_1, 6: num_1, 7: num_1,              # Sample primary temps
        8: num_1, 9: num_1,                         # Sample secondary temps
        10: num_2, 11: num_2, 12: num_1, 13: num_1, # Flows
        14: num_1, 15: num_1, 16: num_2, 17: num_1, # CW
        18: num_1, 19: num_1, 20: num_2,            # CHW
        21: num_0, 22: num_0, 23: num_0,            # Q primary
        24: num_0, 25: num_0,                        # Q secondary, total
        26: num_2,                                   # Tons
        27: num_3, 28: num_3, 29: num_3,            # Area primary
        30: num_3, 31: num_3,                        # Area secondary
    }
    for col_idx, fmt in fmt_map.items():
        for row_idx in range(2, len(feas_results) + 2):
            ws2.cell(row=row_idx, column=col_idx).number_format = fmt

    # Totals row
    tot_row = len(feas_results) + 2
    ws2.cell(row=tot_row, column=1, value="TOTAL").font = tot_font
    ws2.cell(row=tot_row, column=21, value=total_Q_primary).number_format = num_0
    ws2.cell(row=tot_row, column=22, value=total_Q_primary).number_format = num_0
    ws2.cell(row=tot_row, column=24, value=total_Q_secondary).number_format = num_0
    ws2.cell(row=tot_row, column=25, value=total_Q).number_format = num_0
    ws2.cell(row=tot_row, column=26, value=total_tons).number_format = num_2
    for col in range(1, ncols + 1):
        c = ws2.cell(row=tot_row, column=col)
        c.fill = tot_fill
        c.font = tot_font
        c.border = thin_border

    # Freeze header and first column
    ws2.freeze_panes = "B2"
    _auto_width(ws2, ncols, min_width=10, max_width=22)
    # Override a few key columns
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions[get_column_letter(ncols)].width = 50  # Flags

    # ==================================================================
    # SHEET 3 — STEAM ZONE DETAIL
    # ==================================================================
    ws3 = wb.create_sheet("Steam Zones")
    ws3.sheet_properties.tabColor = "BF8F00"

    z_headers = [
        "Sample Point",
        "Zone",
        "Q (Btu/hr)",
        "LMTD (F)",
        "U (Btu/hr-ft2-F)",
        "UA (Btu/hr-F)",
        "A_required (ft2)",
        "Note",
    ]
    z_ncols = len(z_headers)
    for col, h in enumerate(z_headers, 1):
        ws3.cell(row=1, column=col, value=h)
    _style_header(ws3, 1, z_ncols)

    z_row = 2
    for fr in feas_results:
        if not fr.steam_zones_adjusted:
            continue
        for z in fr.steam_zones_adjusted:
            ws3.cell(row=z_row, column=1, value=fr.name)
            ws3.cell(row=z_row, column=2, value=z.zone_name)
            ws3.cell(row=z_row, column=3, value=z.Q_BTU_hr).number_format = num_0
            ws3.cell(row=z_row, column=4, value=z.lmtd_F).number_format = num_1
            ws3.cell(row=z_row, column=5, value=z.U_assumed).number_format = num_0
            ws3.cell(row=z_row, column=6, value=z.UA_required).number_format = num_0
            ws3.cell(row=z_row, column=7, value=z.A_required_ft2).number_format = num_3
            ws3.cell(row=z_row, column=8, value=z.note if z.note else "")
            _style_row(ws3, z_row, z_ncols)
            z_row += 1
        # Sub-total per point
        zones = fr.steam_zones_adjusted
        ws3.cell(row=z_row, column=1, value=fr.name)
        ws3.cell(row=z_row, column=2, value="TOTAL").font = tot_font
        ws3.cell(row=z_row, column=3,
                 value=sum(z.Q_BTU_hr for z in zones)).number_format = num_0
        ws3.cell(row=z_row, column=7,
                 value=sum(z.A_required_ft2 for z in zones
                           if z.A_required_ft2 is not None)).number_format = num_3
        _style_row(ws3, z_row, z_ncols, fill=tot_fill)
        z_row += 1

    _auto_width(ws3, z_ncols, min_width=12, max_width=25)
    ws3.column_dimensions["A"].width = 20

    # ==================================================================
    # SAVE
    # ==================================================================
    wb.save(output_path)
    print(f"  Excel report exported to: {Path(output_path).resolve()}")
    print()
